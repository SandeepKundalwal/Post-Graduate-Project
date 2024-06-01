import os
from openpyxl import load_workbook, Workbook


def copy_data(source_file, destination_sheet, include_headers):
    source_workbook = load_workbook(source_file)
    source_sheet = source_workbook.active

    # Copy column headers only if include_headers is True and it hasn't been copied yet
    if include_headers:
        header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_row += ("ques_no",)  # Add "question_1" column header
        destination_sheet.append(header_row)

    # Copy data rows and append the filename (without extension) to "question_1" column
    filename = os.path.splitext(os.path.basename(source_file))[0]
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        row += (filename,)
        destination_sheet.append(row)

    source_workbook.close()

def merge_xlsx_files(source_folder, destination_file, include_headers):
    # Create a new workbook for the destination
    destination_workbook = Workbook()
    destination_sheet = destination_workbook.active

    # Iterate through all .xlsx files in the source folder
    for filename in os.listdir(source_folder):
        if filename.endswith(".xlsx"):
            print(filename)
            source_file = os.path.join(source_folder, filename)
            copy_data(source_file, destination_sheet, include_headers)
            include_headers = False

    # Save the merged data to the destination file
    destination_workbook.save(destination_file)


if __name__ == "__main__":
    # Replace 'source_folder' and 'destination_file' with your actual folder and file paths
    # C:\Users\DELL\Desktop\Data-Collection\Analysis\Rounds\Human\1
    include_headers = True
    roll_no = "T23023"

    print("\n-------------Round 1--------------------\n")

    correct_round1 = roll_no + "-round1-correct-merged.xlsx"
    incorrect_round1 = roll_no + "-round1-incorrect-merged.xlsx"

    correct_source_folder1 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\1\\correct\\" + roll_no + "\\";
    incorrect_source_folder1 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\1\\incorrect\\" + roll_no + "\\";

    correct_destination_file1 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\1\\standard\\correct\\" + correct_round1;
    incorrect_destination_file1 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\1\\standard\\incorrect\\" + incorrect_round1;

    merge_xlsx_files(correct_source_folder1, correct_destination_file1, include_headers)
    merge_xlsx_files(incorrect_source_folder1, incorrect_destination_file1, include_headers)

    print("\n-------------Round 2--------------------\n")

    correct_round2 = roll_no + "-round2-correct-merged.xlsx"
    incorrect_round2 = roll_no + "-round2-incorrect-merged.xlsx"

    correct_source_folder2 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\2\\correct\\" + roll_no + "\\";
    incorrect_source_folder2 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\2\\incorrect\\" + roll_no + "\\";

    correct_destination_file2 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\2\\standard\\correct\\" + correct_round2;
    incorrect_destination_file2 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\2\\standard\\incorrect\\" + incorrect_round2;

    merge_xlsx_files(correct_source_folder2, correct_destination_file2, include_headers)
    merge_xlsx_files(incorrect_source_folder2, incorrect_destination_file2, include_headers)

    print("\n-------------Round 3--------------------\n")

    correct_round3 = roll_no + "-round3-correct-merged.xlsx"
    incorrect_round3 = roll_no + "-round3-incorrect-merged.xlsx"

    correct_source_folder3 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\3\\correct\\" + roll_no + "\\";
    incorrect_source_folder3 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\3\\incorrect\\" + roll_no + "\\";

    correct_destination_folder3 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\3\\standard\\correct\\" + correct_round3;
    incorrect_destination_file3 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Rounds\\GPT\\3\\standard\\incorrect\\" + incorrect_round3;


    merge_xlsx_files(correct_source_folder3, correct_destination_folder3, include_headers)
    merge_xlsx_files(incorrect_source_folder3, incorrect_destination_file3, include_headers)

    

    # source_folder2 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Genuine-Phishing\\GPT3\\" + roll_no + "\\phishing"
    # destination_file2 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Analysis\\Genuine-Phishing\\GPT3\\" + roll_no + "\\phishing\\" + roll_no + "-phishing-merged.xlsx"

    # merge_xlsx_files(source_folder2, destination_file2, include_headers)