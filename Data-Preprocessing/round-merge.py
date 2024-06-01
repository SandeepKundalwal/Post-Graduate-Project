import os
from openpyxl import load_workbook, Workbook


def copy_data(source_file, destination_sheet, include_headers):
    source_workbook = load_workbook(source_file)
    source_sheet = source_workbook.active

    # Copy column headers only if include_headers is True and it hasn't been copied yet
    if include_headers:
        header_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_row += ("ques_no",)  # Add "question_1" column header
        destination_sheet.append(header_row)

    # Copy data rows and append the filename (without extension) to "question_1" column
    filename = os.path.splitext(os.path.basename(source_file))[0]
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        row += (filename,)
        destination_sheet.append(row)

    source_workbook.close()

def merge_xlsx_files(source_folder, destination_file, include_headers):
    # Create a new workbook for the destination
    destination_workbook = Workbook()
    destination_sheet = destination_workbook.active

    # Iterate through all .xlsx files in the source folder
    for filename in os.listdir(source_folder):
        if filename.endswith(".xlsx"):
            print(filename)
            source_file = os.path.join(source_folder, filename)
            copy_data(source_file, destination_sheet, include_headers)
            include_headers = False

    # Save the merged data to the destination file
    destination_workbook.save(destination_file)


if __name__ == "__main__":
    # Replace 'source_folder' and 'destination_file' with your actual folder and file paths
    include_headers = True
    source_folder = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Rounds\\GPT\\3\\correct\\ayush"
    destination_file = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Rounds\\GPT\\3\\correct\\ayush\\round3-correct-ayush-merged.xlsx"
    source_folder1 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Rounds\\GPT\\3\\incorrect\\ayush"
    destination_file1 = "C:\\Users\\DELL\\Desktop\\Data-Collection\\Rounds\\GPT\\3\\incorrect\\ayush\\round3-incorrect-ayush-merged.xlsx"

    merge_xlsx_files(source_folder, destination_file, include_headers)
    merge_xlsx_files(source_folder1, destination_file1, include_headers)
